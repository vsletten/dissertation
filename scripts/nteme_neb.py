#!/usr/bin/env python3
"""Reconstruct and run the Nteme et al. (2022) muscovite NEB model.

The published supplementary workbook is intentionally kept outside git. This
script extracts its 0.5 K / 0 kbar structure and barrier catalogue, emits a
LAMMPS ``atom_style full`` data file, and prepares one vacancy-hop calculation.

Runtime dependency: openpyxl (for example ``uv run --with openpyxl``).
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import pathlib
import statistics
import sys
from collections import Counter
from dataclasses import asdict, dataclass
from typing import Iterable

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - exercised by CLI users without the extra
    load_workbook = None

WORKBOOK_SHA256 = "292c28c2b182e45aa7fe6cc36a2e62779d7d8dbf0aa6e74525c0e898ee31237f"
EXPECTED_ATOMS = 1512
EXPECTED_COUNTS = {
    "K": 72,
    "Al1": 144,
    "Al2": 72,
    "Si": 216,
    "O1": 144,
    "O2": 432,
    "O3": 288,
    "H": 144,
}

# ClayFF values reported by Nteme et al. (2022), table 1.  Al1 is
# octahedral Al; Al2 is tetrahedral Al.  Energies are kcal/mol.
TYPE_ORDER = ("K", "Al1", "Al2", "Si", "O1", "O2", "O3", "H", "Ar")
MASS = {
    "K": 39.0983,
    "Al1": 26.9815385,
    "Al2": 26.9815385,
    "Si": 28.0855,
    "O1": 15.9994,
    "O2": 15.9994,
    "O3": 15.9994,
    "H": 1.00794,
    "Ar": 39.948,
}
CHARGE = {
    "K": 1.0,
    "Al1": 1.575,
    "Al2": 1.575,
    "Si": 2.1,
    "O1": -0.95,
    "O2": -1.05,
    "O3": -1.16875,
    "H": 0.425,
    "Ar": 0.0,
}
SIGMA = {
    "K": 3.3340,
    "Al1": 4.2712,
    "Al2": 3.3020,
    "Si": 3.3020,
    "O1": 3.1655,
    "O2": 3.1655,
    "O3": 3.1655,
    "H": 0.0,
    "Ar": 3.4010,
}
EPSILON = {
    "K": 0.1000,
    "Al1": 1.33e-6,
    "Al2": 1.84e-6,
    "Si": 1.84e-6,
    "O1": 0.1554,
    "O2": 0.1554,
    "O3": 0.1554,
    "H": 0.0,
    "Ar": 0.2304,
}


@dataclass(frozen=True)
class Atom:
    id: int
    element: str
    x: float
    y: float
    z: float


@dataclass(frozen=True)
class Cell:
    a: float
    b: float
    c: float
    alpha: float
    beta: float
    gamma: float

    def restricted(self) -> tuple[float, float, float, float, float, float]:
        """Return LAMMPS restricted-triclinic lx,ly,lz,xy,xz,yz."""
        alpha = math.radians(self.alpha)
        beta = math.radians(self.beta)
        gamma = math.radians(self.gamma)
        lx = self.a
        xy = self.b * math.cos(gamma)
        xz = self.c * math.cos(beta)
        ly = math.sqrt(self.b * self.b - xy * xy)
        yz = (self.b * self.c * math.cos(alpha) - xy * xz) / ly
        lz2 = self.c * self.c - xz * xz - yz * yz
        if lz2 <= 0:
            raise ValueError("invalid triclinic cell")
        return lx, ly, math.sqrt(lz2), xy, xz, yz


@dataclass(frozen=True)
class Barrier:
    mechanism: str
    species: str
    moving_site: int
    vacancy_1: int
    vacancy_2: int | None
    barrier_kcal_mol: float
    electrostatic_kcal_mol: float | None


def sha256(path: pathlib.Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def require_openpyxl() -> None:
    if load_workbook is None:
        raise SystemExit(
            "openpyxl is required; run with: uv run --with openpyxl python "
            "scripts/nteme_neb.py ..."
        )


def read_workbook(path: pathlib.Path) -> tuple[list[Atom], Cell, list[Barrier]]:
    require_openpyxl()
    actual = sha256(path)
    if actual != WORKBOOK_SHA256:
        raise ValueError(
            f"supplement hash mismatch: expected {WORKBOOK_SHA256}, got {actual}"
        )
    workbook = load_workbook(path, data_only=True, read_only=True)
    structure = workbook["NEB results"]
    atoms: list[Atom] = []
    for row in structure.iter_rows(
        min_row=5, max_row=1516, max_col=5, values_only=True
    ):
        site_id, element, x, y, z = row
        if site_id is not None:
            atoms.append(Atom(int(site_id), str(element), float(x), float(y), float(z)))
    validate_atoms(atoms)

    npt = workbook["NPT run (0.5 K, 0 kbar)"]
    last = tuple(cell.value for cell in npt[253])
    cell = Cell(
        a=float(last[5]),
        b=float(last[6]),
        c=float(last[7]),
        alpha=float(last[8]),
        beta=float(last[9]),
        gamma=float(last[10]),
    )

    barriers: list[Barrier] = []
    barrier_columns = (
        ("monovacancy", "Ar", 9, 10, None, 11, None),
        ("divacancy", "Ar", 14, 15, 16, 17, 18),
        ("monovacancy", "K", 22, 23, None, 24, None),
        ("divacancy", "K", 27, 28, 29, 30, None),
    )
    for row in structure.iter_rows(min_row=5, values_only=True):
        for (
            mechanism,
            species,
            moving,
            vacancy1,
            vacancy2,
            value,
            electrostatic,
        ) in barrier_columns:
            if row[moving - 1] is None or row[value - 1] is None:
                continue
            barriers.append(
                Barrier(
                    mechanism=mechanism,
                    species=species,
                    moving_site=int(row[moving - 1]),
                    vacancy_1=int(row[vacancy1 - 1]),
                    vacancy_2=(
                        int(row[vacancy2 - 1])
                        if vacancy2 and row[vacancy2 - 1] is not None
                        else None
                    ),
                    barrier_kcal_mol=float(row[value - 1]),
                    electrostatic_kcal_mol=(
                        float(row[electrostatic - 1])
                        if electrostatic and row[electrostatic - 1] is not None
                        else None
                    ),
                )
            )
    return atoms, cell, barriers


def validate_atoms(atoms: Iterable[Atom]) -> None:
    atoms = list(atoms)
    if len(atoms) != EXPECTED_ATOMS:
        raise ValueError(f"expected {EXPECTED_ATOMS} atoms, got {len(atoms)}")
    ids = [atom.id for atom in atoms]
    if ids != list(range(1, EXPECTED_ATOMS + 1)):
        raise ValueError("site IDs are not contiguous 1..1512")
    counts = Counter(atom.element for atom in atoms)
    if counts != Counter(EXPECTED_COUNTS):
        raise ValueError(f"unexpected atom counts: {dict(counts)}")
    total_charge = sum(CHARGE[atom.element] for atom in atoms)
    if not math.isclose(total_charge, 0.0, abs_tol=1.0e-10):
        raise ValueError(f"pristine structure is not neutral: q={total_charge}")


def matrix_inverse(cell: Cell) -> tuple[tuple[float, float, float], ...]:
    lx, ly, lz, xy, xz, yz = cell.restricted()
    # Inverse of upper-triangular H where Cartesian r = H fractional s.
    return (
        (1.0 / lx, -xy / (lx * ly), (xy * yz - xz * ly) / (lx * ly * lz)),
        (0.0, 1.0 / ly, -yz / (ly * lz)),
        (0.0, 0.0, 1.0 / lz),
    )


def minimum_image_delta(a: Atom, b: Atom, cell: Cell) -> tuple[float, float, float]:
    dx, dy, dz = b.x - a.x, b.y - a.y, b.z - a.z
    inv = matrix_inverse(cell)
    sx = inv[0][0] * dx + inv[0][1] * dy + inv[0][2] * dz
    sy = inv[1][1] * dy + inv[1][2] * dz
    sz = inv[2][2] * dz
    sx -= round(sx)
    sy -= round(sy)
    sz -= round(sz)
    lx, ly, lz, xy, xz, yz = cell.restricted()
    return lx * sx + xy * sy + xz * sz, ly * sy + yz * sz, lz * sz


def oh_bonds(atoms: list[Atom], cell: Cell) -> list[tuple[int, int]]:
    oxygens = [atom for atom in atoms if atom.element == "O1"]
    hydrogens = [atom for atom in atoms if atom.element == "H"]
    candidates: list[tuple[float, int, int]] = []
    for oxygen in oxygens:
        for hydrogen in hydrogens:
            delta = minimum_image_delta(oxygen, hydrogen, cell)
            distance = math.sqrt(sum(value * value for value in delta))
            if distance < 1.35:
                candidates.append((distance, oxygen.id, hydrogen.id))
    candidates.sort()
    used_o: set[int] = set()
    used_h: set[int] = set()
    bonds: list[tuple[int, int]] = []
    for distance, oxygen_id, hydrogen_id in candidates:
        if oxygen_id not in used_o and hydrogen_id not in used_h:
            used_o.add(oxygen_id)
            used_h.add(hydrogen_id)
            bonds.append((oxygen_id, hydrogen_id))
    if len(bonds) != 144 or len(used_o) != 144 or len(used_h) != 144:
        raise ValueError(
            f"failed to identify 144 one-to-one O1-H bonds; found {len(bonds)}"
        )
    return sorted(bonds)


def lammps_potential_lines() -> list[str]:
    lines = [
        "pair_style      lj/cut/coul/long 10.0",
        "pair_modify     mix arithmetic",
    ]
    for index, name in enumerate(TYPE_ORDER, 1):
        lines.append(
            f"pair_coeff      {index} {index} {EPSILON[name]:.10g} {SIGMA[name]:.10g}"
        )
    lines.extend(
        (
            "bond_style      harmonic",
            "bond_coeff      1 554.1349 1.0",
            "special_bonds   lj/coul 0.0 0.0 0.0",
            "kspace_style    ewald 1.0e-4",
            "neighbor        2.0 bin",
            "neigh_modify    delay 0 every 1 check yes",
        )
    )
    return lines


def write_data(path: pathlib.Path, atoms: list[Atom], cell: Cell) -> None:
    bonds = oh_bonds(atoms, cell)
    lx, ly, lz, xy, xz, yz = cell.restricted()
    type_id = {name: index for index, name in enumerate(TYPE_ORDER, 1)}
    lines = [
        "Nteme et al. (2022) 6x3x1 muscovite, reconstructed from supplement",
        "",
        f"{len(atoms)} atoms",
        f"{len(bonds)} bonds",
        "",
        f"{len(TYPE_ORDER)} atom types",
        "1 bond types",
        "",
        f"0.0 {lx:.12f} xlo xhi",
        f"0.0 {ly:.12f} ylo yhi",
        f"0.0 {lz:.12f} zlo zhi",
        f"{xy:.12f} {xz:.12f} {yz:.12f} xy xz yz",
        "",
        "Masses",
        "",
    ]
    for index, name in enumerate(TYPE_ORDER, 1):
        lines.append(f"{index} {MASS[name]:.10f} # {name}")
    lines.extend(("", "Atoms # full", ""))
    for atom in atoms:
        lines.append(
            f"{atom.id} 1 {type_id[atom.element]} {CHARGE[atom.element]:.8f} "
            f"{atom.x:.12f} {atom.y:.12f} {atom.z:.12f} # {atom.element}"
        )
    lines.extend(("", "Bonds", ""))
    for bond_id, (oxygen_id, hydrogen_id) in enumerate(bonds, 1):
        lines.append(f"{bond_id} 1 {oxygen_id} {hydrogen_id}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_static_input(path: pathlib.Path, data_name: str) -> None:
    lines = [
        "units           real",
        "atom_style      full",
        "atom_modify     map array sort 0 0.0",
        "boundary        p p p",
        f"read_data       {data_name}",
        *lammps_potential_lines(),
        "thermo_style    custom step atoms bonds pe ebond evdwl ecoul elong press",
        "thermo          1",
        "run             0",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_neb_input(
    path: pathlib.Path,
    data_name: str,
    final_name: str,
    *,
    etol: float,
    ftol: float,
    relax_steps: int,
    climb_steps: int,
) -> None:
    lines = [
        "units           real",
        "atom_style      full",
        "atom_modify     map array sort 0 0.0",
        "boundary        p p p",
        f"read_data       {data_name}",
        *lammps_potential_lines(),
        "thermo_style    custom step pe ebond evdwl ecoul elong fnorm fmax",
        "thermo          100",
        "timestep        0.1",
        "min_style       quickmin",
        "variable        u uloop 8",
        "fix             path all neb 23.0",
        "dump            path all custom 100 dump.neb.$u id type x y z",
        f"neb             {etol:.8g} {ftol:.8g} {relax_steps} {climb_steps} 100 final {final_name}",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def prepare_route(
    atoms: list[Atom],
    cell: Cell,
    route: Barrier,
) -> tuple[list[Atom], tuple[float, float, float]]:
    """Build the initial defective cell and final moving-atom coordinate."""
    by_id = {atom.id: atom for atom in atoms}
    moving = by_id[route.moving_site]
    destination = by_id[route.vacancy_1]
    if moving.element != "K" or destination.element != "K":
        raise ValueError("published routes must connect K-gallery sites")
    if route.species != "Ar":
        raise ValueError("only published Ar hops are currently executable")
    removed = {route.vacancy_1}
    if route.vacancy_2 is not None:
        removed.add(route.vacancy_2)
    prepared: list[Atom] = []
    for atom in atoms:
        if atom.id in removed:
            continue
        if atom.id == route.moving_site:
            prepared.append(Atom(atom.id, "Ar", atom.x, atom.y, atom.z))
        else:
            prepared.append(atom)
    dx, dy, dz = minimum_image_delta(moving, destination, cell)
    final = moving.x + dx, moving.y + dy, moving.z + dz
    return prepared, final


def select_route(barriers: list[Barrier], mechanism: str, route_number: int) -> Barrier:
    candidates = [
        route
        for route in barriers
        if route.species == "Ar" and route.mechanism == mechanism
    ]
    if not 1 <= route_number <= len(candidates):
        raise ValueError(
            f"route number must be in 1..{len(candidates)} for Ar {mechanism}"
        )
    return candidates[route_number - 1]


def write_catalogue(path: pathlib.Path, barriers: list[Barrier]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(asdict(barriers[0])))
        writer.writeheader()
        for barrier in barriers:
            writer.writerow(asdict(barrier))


def command_extract(args: argparse.Namespace) -> None:
    atoms, cell, barriers = read_workbook(args.workbook)
    args.out.mkdir(parents=True, exist_ok=True)
    write_data(args.out / "pristine.data", atoms, cell)
    write_static_input(args.out / "in.static", "pristine.data")
    write_catalogue(args.out / "published-barriers.csv", barriers)
    barrier_stats = {}
    for species in ("Ar", "K"):
        for mechanism in ("monovacancy", "divacancy"):
            values = [
                barrier.barrier_kcal_mol
                for barrier in barriers
                if barrier.species == species and barrier.mechanism == mechanism
            ]
            barrier_stats[f"{species}-{mechanism}"] = {
                "n": len(values),
                "mean_kcal_mol": statistics.mean(values),
                "stdev_kcal_mol": statistics.stdev(values),
                "min_kcal_mol": min(values),
                "max_kcal_mol": max(values),
            }
    summary = {
        "source": str(args.workbook.resolve()),
        "source_sha256": sha256(args.workbook),
        "atoms": len(atoms),
        "atom_counts": dict(Counter(atom.element for atom in atoms)),
        "cell": asdict(cell),
        "restricted_cell": dict(
            zip(("lx", "ly", "lz", "xy", "xz", "yz"), cell.restricted())
        ),
        "oh_bonds": len(oh_bonds(atoms, cell)),
        "barrier_rows": len(barriers),
        "barrier_counts": dict(Counter(f"{b.species}-{b.mechanism}" for b in barriers)),
        "barrier_stats": barrier_stats,
    }
    (args.out / "source-summary.json").write_text(
        json.dumps(summary, indent=2) + "\n", encoding="utf-8"
    )
    print(json.dumps(summary, indent=2))


def command_prepare(args: argparse.Namespace) -> None:
    atoms, cell, barriers = read_workbook(args.workbook)
    route = select_route(barriers, args.mechanism, args.route_number)
    prepared, final = prepare_route(atoms, cell, route)
    args.out.mkdir(parents=True, exist_ok=True)
    write_data(args.out / "initial.data", prepared, cell)
    (args.out / "final.coords").write_text(
        f"1\n{route.moving_site} {final[0]:.12f} {final[1]:.12f} {final[2]:.12f}\n",
        encoding="utf-8",
    )
    write_neb_input(
        args.out / "in.neb",
        "initial.data",
        "final.coords",
        etol=args.etol,
        ftol=args.ftol,
        relax_steps=args.relax_steps,
        climb_steps=args.climb_steps,
    )
    manifest = {
        "source": str(args.workbook.resolve()),
        "source_sha256": sha256(args.workbook),
        "route_number": args.route_number,
        "route": asdict(route),
        "prepared_atoms": len(prepared),
        "net_charge_e": sum(CHARGE[atom.element] for atom in prepared),
        "final_moving_coordinate_angstrom": final,
        "images": 8,
        "spring_kcal_mol_angstrom2": 23.0,
        "etol": args.etol,
        "ftol_kcal_mol_angstrom": args.ftol,
        "relax_steps": args.relax_steps,
        "climb_steps": args.climb_steps,
        "run_command": "mpirun -np 8 lmp -partition 8x1 -in in.neb",
        "known_replication_limit": (
            "The paper and supplement do not identify the remote tetrahedral "
            "charge-compensation substitutions for each route. This reconstruction "
            "therefore preserves the published defective-cell net charge and LAMMPS "
            "Ewald neutralizing-background behavior."
        ),
    }
    (args.out / "run-manifest.json").write_text(
        json.dumps(manifest, indent=2) + "\n", encoding="utf-8"
    )
    print(json.dumps(manifest, indent=2))


def command_summarize(args: argparse.Namespace) -> None:
    manifest_path = args.run_dir / "run-manifest.json"
    screen_path = args.run_dir / "screen.log"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    numeric_rows: list[list[float]] = []
    for line in screen_path.read_text(encoding="utf-8").splitlines():
        fields = line.split()
        if len(fields) < 10:
            continue
        try:
            numeric_rows.append([float(field) for field in fields])
        except ValueError:
            continue
    if not numeric_rows:
        raise ValueError(f"no NEB progress rows found in {screen_path}")
    row = numeric_rows[-1]
    if len(row) != 25:
        raise ValueError(f"expected 25 NEB columns, got {len(row)}")
    published = float(manifest["route"]["barrier_kcal_mol"])
    computed = row[6]
    signed_error = computed - published
    final_step = int(row[0])
    recent_barriers = [
        progress[6] for progress in numeric_rows if int(progress[0]) >= final_step - 500
    ]
    result = {
        "route": manifest["route"],
        "final_step": final_step,
        "max_replica_force_kcal_mol_angstrom": row[1],
        "max_atom_force_kcal_mol_angstrom": row[2],
        "computed_forward_barrier_kcal_mol": computed,
        "computed_reverse_barrier_kcal_mol": row[7],
        "published_route_barrier_kcal_mol": published,
        "signed_error_kcal_mol": signed_error,
        "absolute_error_kcal_mol": abs(signed_error),
        "signed_percent_error": 100.0 * signed_error / published,
        "barrier_span_last_500_steps_kcal_mol": max(recent_barriers)
        - min(recent_barriers),
        "converged_to_requested_ftol": row[1]
        <= float(manifest["ftol_kcal_mol_angstrom"]),
        "screen_sha256": sha256(screen_path),
    }
    output_path = args.run_dir / "result-summary.json"
    output_path.write_text(json.dumps(result, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(result, indent=2))


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    sub = parser.add_subparsers(required=True)
    extract = sub.add_parser(
        "extract", help="extract the published structure and barrier catalogue"
    )
    extract.add_argument("--workbook", type=pathlib.Path, required=True)
    extract.add_argument("--out", type=pathlib.Path, required=True)
    extract.set_defaults(func=command_extract)
    prepare = sub.add_parser(
        "prepare", help="prepare one published Ar vacancy hop for LAMMPS NEB"
    )
    prepare.add_argument("--workbook", type=pathlib.Path, required=True)
    prepare.add_argument("--out", type=pathlib.Path, required=True)
    prepare.add_argument(
        "--mechanism", choices=("monovacancy", "divacancy"), required=True
    )
    prepare.add_argument("--route-number", type=int, default=1)
    prepare.add_argument("--etol", type=float, default=1.0e-10)
    prepare.add_argument("--ftol", type=float, default=1.0e-4)
    prepare.add_argument("--relax-steps", type=int, default=2000)
    prepare.add_argument("--climb-steps", type=int, default=5000)
    prepare.set_defaults(func=command_prepare)
    summarize = sub.add_parser(
        "summarize", help="summarize a completed LAMMPS NEB screen log"
    )
    summarize.add_argument("--run-dir", type=pathlib.Path, required=True)
    summarize.set_defaults(func=command_summarize)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    args.func(args)
    return 0


if __name__ == "__main__":
    sys.exit(main())
